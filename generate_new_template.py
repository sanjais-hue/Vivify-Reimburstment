import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Reimbursement"

# Headers
headers = [
    "Serial Number", "Date", "Expense Type", "Details / Particulars", 
    "Conveyance Amt", "Lodging Amt", "Fooding Amt", "Others Amt", "Misc Amt", 
    "Total Amount",
    "SMO/WBS No", "SO/SAP No", "Ref No", "Remarks",
    "From Time", "To Time", "Distance (KM)", "Mode of Trans"
]

# Style for header
header_fill = PatternFill(start_color="3f418d", end_color="3f418d", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = border
    
    # Adjust column width
    if "Amt" in header or "Amount" in header:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15
    else:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

# Add sample row with formulas
# Columns E to I are amount columns. Total is column J (10).
# SUM(E2:I2)
sample_rows = [
    ["1", "2026-02-26", "Local", "Local Visit", 150.00, 0, 0, 0, 0, "=SUM(E2:I2)", "SMO-100", "SO-500", "REF-01", "Met client", "09:00 AM", "11:00 AM", "15", "Bike"],
    ["2", "2026-02-26", "Tour", "Hotel & Food", 0, 2500.00, 450.00, 0, 0, "=SUM(E3:I3)", "SMO-100", "SO-500", "REF-01", "Site stay", "", "", "", ""],
]

for row_idx, row_data in enumerate(sample_rows, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        if isinstance(value, (int, float)):
             cell.number_format = '#,##0.00'

wb.save("Reimbursement_Template.xlsx")
print("New Template generated successfully as Reimbursement_Template.xlsx")
